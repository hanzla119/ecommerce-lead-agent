import datetime
from pathlib import Path
from typing import List, Dict
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from config import OUTPUT_DIR

def transform_lead_to_row(lead: Dict) -> Dict[str, any]:
    """Flattens nested lead dictionary into a clean spreadsheet row."""
    contacts = lead.get("contacts", {})
    audit = lead.get("audit", {})
    pitches = lead.get("pitches", {})
    
    return {
        "Brand Name": contacts.get("brand_name", "") or lead.get("url", ""),
        "Store URL": lead.get("url", ""),
        "Region": lead.get("region", "UK"),
        "Contact Email": contacts.get("email", ""),
        "Email Deliverability": contacts.get("email_deliverability", "Unverified"),
        "Phone": contacts.get("phone", ""),
        "Founder / Decision Maker": contacts.get("founder_name", ""),
        "Founder Title": contacts.get("founder_title", ""),
        "Founder LinkedIn Profile": contacts.get("founder_linkedin", ""),
        "Instagram Handle": contacts.get("instagram_handle", ""),
        "Instagram URL": contacts.get("instagram", ""),
        "Facebook URL": contacts.get("facebook", ""),
        "TikTok Handle": contacts.get("tiktok_handle", ""),
        "Platform": audit.get("platform", "Shopify"),
        "Response Speed (ms)": audit.get("response_time_ms", 0),
        "Meta Pixel Active": "ACTIVE ✅" if audit.get("has_meta_pixel") else "MISSING ❌",
        "TikTok Pixel Active": "ACTIVE ✅" if audit.get("has_tiktok_pixel") else "MISSING ❌",
        "Google Analytics": "ACTIVE ✅" if audit.get("has_google_analytics") else "MISSING ❌",
        "Email Flow (Klaviyo)": "ACTIVE ✅" if audit.get("has_klaviyo") else "MISSING ❌",
        "Reviews App": audit.get("reviews_app_name", "None"),
        "Critical Leaks Detected": " | ".join(audit.get("critical_gaps", [])),
        "Email Subject 1 (Scale Proof)": pitches.get("email_subject", ""),
        "Email Body 1 (Scale Proof)": pitches.get("email_body", ""),
        "Email Subject 2 (Tracking Leak)": pitches.get("email_subject_alt", ""),
        "Email Body 2 (Tracking Leak)": pitches.get("email_body_alt", ""),
        "Email Subject 3 (CRO 4.89%)": pitches.get("email_subject_cro", ""),
        "Email Body 3 (CRO 4.89%)": pitches.get("email_body_cro", ""),
        "Email Subject 4 (Short Hook)": pitches.get("email_subject_short", ""),
        "Email Body 4 (Short Hook)": pitches.get("email_body_short", ""),
        "LinkedIn Connection Note": pitches.get("linkedin_note", ""),
        "Instagram Direct Message": pitches.get("instagram_dm", ""),
        "Outreach Status": "New Lead",
        "Date Discovered": datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    }

def export_leads_to_files(leads: List[Dict], filename_prefix: str = "leads") -> Dict[str, str]:
    """
    Exports leads list to both timestamped CSV/Excel and updates the latest master files.
    """
    if not leads:
        return {}
        
    try:
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    except Exception:
        pass
        
    rows = [transform_lead_to_row(l) for l in leads]
    df = pd.DataFrame(rows)
    
    # Check if latest CSV already exists to append/merge
    latest_csv = OUTPUT_DIR / "leads_latest.csv"
    latest_excel = OUTPUT_DIR / "leads_latest.xlsx"
    
    if latest_csv.exists():
        try:
            existing_df = pd.read_csv(latest_csv, encoding="utf-8-sig")
            # Deduplicate based on Store URL
            combined_df = pd.concat([df, existing_df], ignore_index=True)
            combined_df = combined_df.drop_duplicates(subset=["Store URL"], keep="first")
            df = combined_df
        except Exception:
            pass

    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    timestamp_csv = OUTPUT_DIR / f"{filename_prefix}_{timestamp}.csv"
    timestamp_excel = OUTPUT_DIR / f"{filename_prefix}_{timestamp}.xlsx"
    
    # Save CSVs
    try:
        df.to_csv(latest_csv, index=False, encoding="utf-8-sig")
        df.to_csv(timestamp_csv, index=False, encoding="utf-8-sig")
    except Exception as e:
        print(f"CSV export warning: {e}")
        
    # Save Styled Excel
    try:
        with pd.ExcelWriter(latest_excel, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Qualified Leads")
            workbook = writer.book
            worksheet = writer.sheets["Qualified Leads"]
            
            # Styling Headers
            header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            
            for col_idx, col_name in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                worksheet.column_dimensions[cell.column_letter].width = max(len(col_name) + 5, 18)
                
            worksheet.freeze_panes = "A2"
            
        with pd.ExcelWriter(timestamp_excel, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Qualified Leads")
    except Exception as e:
        print(f"Excel export warning: {e}")
        
    return {
        "csv": str(timestamp_csv),
        "excel": str(timestamp_excel),
        "latest_csv": str(latest_csv),
        "latest_excel": str(latest_excel),
        "count": len(df)
    }
